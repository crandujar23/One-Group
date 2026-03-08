from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import logging
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import Http404, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_http_methods

from openpyxl import load_workbook

from core.rbac.constants import ModuleCode, PermissionAction, RoleCode
from core.rbac.services import has_module_permission
from crm.forms import CrmDealExcelUploadForm, CrmDealSalesrepForm
from crm.models import CrmDeal, SalesRep
from crm.serializers import CrmDealDetailSerializer
from dashboard.services.hierarchy_scope_service import get_downline_user_ids

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class DealAccess:
    my_salesrep: SalesRep | None
    can_view: bool
    can_reassign: bool
    can_delete: bool
    is_global: bool
    visible_user_ids: set[int]


def _profile(user):
    return getattr(user, "profile", None)


def _my_salesrep(user) -> SalesRep | None:
    return SalesRep.objects.select_related("user").filter(user=user, is_active=True).first()


def _display_name_from_rep(rep: SalesRep | None) -> str:
    if not rep or not rep.user_id:
        return ""
    return rep.user.get_full_name().strip() or rep.user.get_username()


def _can_access_sales_section(user) -> bool:
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    profile = _profile(user)
    if not profile or not getattr(profile, "role", ""):
        return False
    return bool(
        has_module_permission(user, module=ModuleCode.SALES, action=PermissionAction.VIEW)
        or has_module_permission(user, module=ModuleCode.SALES, action=PermissionAction.MANAGE)
    )


def _deal_access(user) -> DealAccess:
    profile = _profile(user)
    my_rep = _my_salesrep(user)
    is_global = bool(user.is_superuser or (profile and profile.role in {RoleCode.PARTNER, RoleCode.ADMINISTRADOR}))
    visible_ids = get_downline_user_ids(user) if user.is_authenticated else set()
    can_reassign = bool(
        user.is_superuser
        or user.has_perm("crm.can_reassign_deal")
        or (profile and profile.role in {RoleCode.PARTNER, RoleCode.ADMINISTRADOR})
    )
    can_delete = bool(
        user.is_superuser
        or user.has_perm("crm.delete_crmdeal")
        or (profile and profile.role in {RoleCode.PARTNER, RoleCode.ADMINISTRADOR})
    )
    can_view = _can_access_sales_section(user) and (is_global or my_rep is not None)
    return DealAccess(
        my_salesrep=my_rep,
        can_view=can_view,
        can_reassign=can_reassign,
        can_delete=can_delete,
        is_global=is_global,
        visible_user_ids=visible_ids,
    )


def _deals_queryset_for_user(user, *, deal_kind: str):
    access = _deal_access(user)
    if not access.can_view:
        return CrmDeal.objects.none()
    qs = CrmDeal.objects.filter(deal_kind=deal_kind).select_related("salesrep__user", "imported_by")
    if access.is_global:
        return qs
    if access.visible_user_ids and len(access.visible_user_ids) > 1:
        q = Q(salesrep__user_id__in=access.visible_user_ids)
        if access.can_reassign:
            q |= Q(salesrep__isnull=True)
        return qs.filter(q)
    if access.my_salesrep:
        return qs.filter(salesrep=access.my_salesrep)
    return CrmDeal.objects.none()


def _salesrep_choices_for_user(user):
    access = _deal_access(user)
    qs = SalesRep.objects.select_related("user").filter(is_active=True)
    if access.is_global:
        return qs.order_by("user__first_name", "user__last_name", "user__username")
    if access.visible_user_ids:
        return qs.filter(user_id__in=access.visible_user_ids).order_by("user__first_name", "user__last_name", "user__username")
    return SalesRep.objects.none()


def _normalize_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    raw = str(value).strip().replace("$", "").replace(" ", "")
    if not raw:
        return None
    if "," in raw and "." in raw:
        raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


def _normalize_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    parsed = parse_date(raw)
    if parsed:
        return parsed
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _norm_header(value: str) -> str:
    text = re.sub(r"[^A-Z0-9]+", " ", (value or "").upper()).strip()
    return re.sub(r"\s+", " ", text)


def _sync_deal_hierarchy_snapshot(deal: CrmDeal, salesrep: SalesRep | None):
    if not salesrep:
        deal.consultant_name = ""
        deal.advisor_name = ""
        deal.manager_name = ""
        deal.senior_manager_name = ""
        deal.elite_manager_name = ""
        deal.business_manager_name = ""
        deal.jr_partner_name = ""
        deal.partner_name = ""
        deal.consultant_rate = Decimal("0")
        deal.advisor_rate = Decimal("0")
        deal.manager_rate = Decimal("0")
        deal.senior_manager_rate = Decimal("0")
        deal.elite_manager_rate = Decimal("0")
        deal.business_manager_rate = Decimal("0")
        deal.jr_partner_rate = Decimal("0")
        deal.partner_rate = Decimal("0")
        return

    deal.consultant_name = _display_name_from_rep(salesrep)
    deal.advisor_name = _display_name_from_rep(salesrep.consultant)
    deal.manager_name = _display_name_from_rep(salesrep.teamleader)
    deal.senior_manager_name = _display_name_from_rep(salesrep.manager)
    deal.elite_manager_name = _display_name_from_rep(salesrep.promanager)
    deal.business_manager_name = _display_name_from_rep(salesrep.executivemanager)
    deal.jr_partner_name = _display_name_from_rep(salesrep.jr_partner)
    deal.partner_name = _display_name_from_rep(salesrep.partner)
    deal.consultant_rate = salesrep.trainee_rate or Decimal("0")
    deal.advisor_rate = salesrep.consultant_rate or Decimal("0")
    deal.manager_rate = salesrep.teamleader_rate or Decimal("0")
    deal.senior_manager_rate = salesrep.manager_rate or Decimal("0")
    deal.elite_manager_rate = salesrep.promanager_rate or Decimal("0")
    deal.business_manager_rate = salesrep.executivemanager_rate or Decimal("0")
    deal.jr_partner_rate = salesrep.jr_partner_rate or Decimal("0")
    deal.partner_rate = salesrep.partner_rate or Decimal("0")


def _match_salesrep_by_name(name: str) -> SalesRep | None:
    candidate = (name or "").strip()
    if not candidate:
        return None
    qs = SalesRep.objects.select_related("user").filter(is_active=True)
    for rep in qs:
        display = rep.user.get_full_name().strip() or rep.user.get_username()
        if display.lower() == candidate.lower():
            return rep
    return None


def _parse_contract_and_proposal(raw_value: str) -> tuple[str, str]:
    value = (raw_value or "").strip()
    if not value:
        return "", ""
    parts = [part.strip() for part in re.split(r"[/|]+", value) if part.strip()]
    if len(parts) >= 2:
        return parts[0], parts[1]
    return value, value


def _import_deals_from_excel(*, file_obj, sheet_name: str, dry_run: bool, actor, deal_kind: str) -> dict:
    summary = {
        "processed": 0,
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "warnings": [],
        "errors": [],
    }
    wb = load_workbook(file_obj, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        summary["errors"].append("La hoja seleccionada no contiene datos.")
        return summary
    header = [_norm_header(str(cell or "")) for cell in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def col(*names):
        for name in names:
            if name in idx:
                return idx[name]
        return None

    col_contract = col("SERVICE CONTRACT PROPOSAL ID", "SERVICE CONTRACT PROPOSAL ID ")
    col_rep_name = col("SALES REP NAME", "SALES REP NAME ")
    col_date_approved = col("DATE APPROVED")
    col_epc_priced = col("EPC PRICED")
    col_system_size = col("SYSTEM SIZE DC")
    col_epc_base = col("EPC BASE")
    col_epc_table = col("EPC TABLA")
    col_epc_adj = col("AJUSTE POR EPC")

    for row_idx, row in enumerate(rows[1:], start=2):
        summary["processed"] += 1
        contract_raw = str(row[col_contract] or "").strip() if col_contract is not None and col_contract < len(row) else ""
        if not contract_raw:
            summary["skipped"] += 1
            summary["warnings"].append(f"Fila {row_idx}: sin SERVICE CONTRACT+/PROPOSAL ID.")
            continue

        service_contract, proposal = _parse_contract_and_proposal(contract_raw)
        imported_salesrep_name = str(row[col_rep_name] or "").strip() if col_rep_name is not None and col_rep_name < len(row) else ""

        existing = CrmDeal.objects.filter(deal_kind=deal_kind).filter(
            Q(sunrun_service_contract_id=service_contract) | Q(proposal_id=proposal)
        ).first()
        deal = existing or CrmDeal(deal_kind=deal_kind)

        deal.sunrun_service_contract_id = service_contract
        deal.proposal_id = proposal
        deal.imported_salesrep_name = imported_salesrep_name
        deal.imported_by = actor
        deal.imported_at = timezone.now()
        deal.closing_date = _normalize_date(row[col_date_approved] if col_date_approved is not None and col_date_approved < len(row) else None)
        deal.sr_signoff_date = deal.closing_date
        deal.epc_price = _normalize_decimal(row[col_epc_priced] if col_epc_priced is not None and col_epc_priced < len(row) else None)
        deal.system_size = _normalize_decimal(row[col_system_size] if col_system_size is not None and col_system_size < len(row) else None)
        deal.epc_base = _normalize_decimal(row[col_epc_base] if col_epc_base is not None and col_epc_base < len(row) else None)
        deal.epc_table = _normalize_decimal(row[col_epc_table] if col_epc_table is not None and col_epc_table < len(row) else None)
        deal.epc_adjustment = _normalize_decimal(row[col_epc_adj] if col_epc_adj is not None and col_epc_adj < len(row) else None)
        if not deal.stage:
            deal.stage = CrmDeal.Stage.PLANNED

        matched_rep = _match_salesrep_by_name(imported_salesrep_name)
        if matched_rep:
            deal.salesrep = matched_rep
            _sync_deal_hierarchy_snapshot(deal, matched_rep)
        elif imported_salesrep_name:
            summary["warnings"].append(f'Fila {row_idx}: no se encontro asociado "{imported_salesrep_name}".')

        if dry_run:
            if existing:
                summary["updated"] += 1
            else:
                summary["created"] += 1
            continue

        try:
            deal.save()
            if existing:
                summary["updated"] += 1
            else:
                summary["created"] += 1
        except Exception as exc:
            summary["errors"].append(f"Fila {row_idx}: error guardando registro ({exc}).")

    if len(summary["warnings"]) > 20:
        summary["warnings_hidden"] = len(summary["warnings"]) - 20
        summary["warnings"] = summary["warnings"][:20]
    else:
        summary["warnings_hidden"] = 0
    if len(summary["errors"]) > 20:
        summary["errors_hidden"] = len(summary["errors"]) - 20
        summary["errors"] = summary["errors"][:20]
    else:
        summary["errors_hidden"] = 0
    return summary


def _compute_deal_kpis(qs):
    now = timezone.localdate()
    total = qs.count()
    total_pipeline = qs.aggregate(total=Sum("epc_price"))["total"] or Decimal("0")
    month_total = qs.filter(closing_date__year=now.year, closing_date__month=now.month).count()
    stage_counts = {}
    for stage_key, stage_label in CrmDeal.Stage.choices:
        value = qs.filter(stage=stage_key).count()
        if value:
            stage_counts[stage_label] = value
    return {
        "total": total,
        "total_pipeline": float(total_pipeline),
        "month_total": month_total,
        "stage_counts": stage_counts,
    }


def _month_options(qs):
    values = (
        qs.exclude(closing_date__isnull=True)
        .dates("closing_date", "month", order="DESC")
    )
    return [
        {"value": f"{item.year:04d}-{item.month:02d}", "label": item.strftime("%B %Y")}
        for item in values
    ]


@login_required
@require_http_methods(["GET", "POST"])
def crm_deals_list_page(request):
    access = _deal_access(request.user)
    if not access.can_view:
        return JsonResponse({"detail": "No autorizado"}, status=403)

    deal_kind = (request.GET.get("deal_kind") or request.POST.get("deal_kind") or CrmDeal.DealKind.RESIDENTIAL).strip() or CrmDeal.DealKind.RESIDENTIAL
    base_qs = _deals_queryset_for_user(request.user, deal_kind=deal_kind)
    upload_summary = None
    upload_form = CrmDealExcelUploadForm()

    if request.method == "POST":
        if not access.can_reassign:
            return JsonResponse({"detail": "No autorizado"}, status=403)
        upload_form = CrmDealExcelUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            cleaned = upload_form.cleaned_data
            try:
                upload_summary = _import_deals_from_excel(
                    file_obj=cleaned["report_file"],
                    sheet_name=cleaned.get("sheet_name") or "",
                    dry_run=bool(cleaned.get("dry_run")),
                    actor=request.user,
                    deal_kind=deal_kind,
                )
                if cleaned.get("dry_run"):
                    messages.info(request, "Validacion completada (dry run).")
                else:
                    messages.success(request, "Informe procesado correctamente.")
            except Exception:
                logger.exception("Error procesando importacion de deals")
                messages.error(request, "No fue posible procesar el informe.")

    context = {
        "title": "Pipeline de Deals",
        "deal_kind": deal_kind,
        "upload_form": upload_form,
        "upload_summary": upload_summary,
        "can_reassign_deals_ui": access.can_reassign,
        "can_delete_deals_ui": access.can_delete,
        "api_url": "/apps/api/deals-details/",
        "stages": list(CrmDeal.Stage.choices),
        "month_options": _month_options(base_qs),
        "kpis": _compute_deal_kpis(base_qs),
    }
    return render(request, "dashboard/deals/deals_list.html", context)


@login_required
@require_http_methods(["GET"])
def crm_deals_details_api(request):
    access = _deal_access(request.user)
    if not access.can_view:
        return JsonResponse({"detail": "No autorizado"}, status=403)
    deal_kind = (request.GET.get("deal_kind") or CrmDeal.DealKind.RESIDENTIAL).strip() or CrmDeal.DealKind.RESIDENTIAL
    stage = (request.GET.get("stage") or "").strip()
    month = (request.GET.get("month") or "").strip()
    search = (request.GET.get("search") or request.GET.get("search[value]") or "").strip()

    qs = _deals_queryset_for_user(request.user, deal_kind=deal_kind)
    if stage:
        qs = qs.filter(stage=stage)
    if month and re.match(r"^\d{4}-\d{2}$", month):
        year, mon = month.split("-")
        qs = qs.filter(closing_date__year=int(year), closing_date__month=int(mon))
    if search:
        qs = qs.filter(
            Q(customer_name__icontains=search)
            | Q(customer_email__icontains=search)
            | Q(customer_phone__icontains=search)
            | Q(proposal_id__icontains=search)
            | Q(sunrun_service_contract_id__icontains=search)
            | Q(stage__icontains=search)
            | Q(imported_salesrep_name__icontains=search)
        )

    rows = CrmDealDetailSerializer.serialize_many(qs.order_by("-closing_date", "-id"))
    for row in rows:
        row["can_edit"] = access.can_reassign
        row["can_delete"] = access.can_delete
    return JsonResponse({"data": rows, "kpis": _compute_deal_kpis(qs)})


def _visible_deal_or_404(user, deal_id: int, deal_kind: str) -> CrmDeal:
    qs = _deals_queryset_for_user(user, deal_kind=deal_kind)
    try:
        return qs.get(pk=deal_id)
    except CrmDeal.DoesNotExist as exc:
        raise Http404 from exc


@login_required
@require_http_methods(["GET", "POST"])
def crm_deal_update_modal(request, deal_id: int):
    deal_kind = (request.GET.get("deal_kind") or request.POST.get("deal_kind") or CrmDeal.DealKind.RESIDENTIAL).strip() or CrmDeal.DealKind.RESIDENTIAL
    deal = _visible_deal_or_404(request.user, deal_id=deal_id, deal_kind=deal_kind)
    access = _deal_access(request.user)
    if not access.can_reassign:
        return JsonResponse({"success": False, "message": "No autorizado."}, status=403)

    if request.method == "GET":
        form = CrmDealSalesrepForm(instance=deal, salesrep_queryset=_salesrep_choices_for_user(request.user))
        return render(request, "dashboard/deals/_deal_update_modal.html", {"deal": deal, "form": form, "deal_kind": deal_kind})

    form = CrmDealSalesrepForm(request.POST, instance=deal, salesrep_queryset=_salesrep_choices_for_user(request.user))
    if not form.is_valid():
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    updated_deal = form.save(commit=False)
    _sync_deal_hierarchy_snapshot(updated_deal, updated_deal.salesrep)
    updated_deal.save()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["DELETE"])
def crm_deal_delete_api(request, deal_id: int):
    deal_kind = (request.GET.get("deal_kind") or CrmDeal.DealKind.RESIDENTIAL).strip() or CrmDeal.DealKind.RESIDENTIAL
    deal = _visible_deal_or_404(request.user, deal_id=deal_id, deal_kind=deal_kind)
    access = _deal_access(request.user)
    if not access.can_delete:
        return JsonResponse({"success": False, "message": "No autorizado."}, status=403)
    deal.delete()
    return JsonResponse({"success": True, "message": "Deal eliminado correctamente."})
